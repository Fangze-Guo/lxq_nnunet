import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_subtables_from_excel():
    """
    从Excel文件中创建两个子表格
    """
    
    # 文件路径
    excel_file = "/media/dell/T7 Shield/nnunet/AllData/ZJYY/珠江需要重新导出10例.xlsx"
    
    # 需要筛选的PatientID列表
    patient_ids_to_filter = [
        "4167356", "4600423", "4605329", "4619739", "4646742", 
        "4673571", "4690443", "4938852", "2901533","4128007"
    ]
    
    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误: Excel文件不存在: {excel_file}")
        return False
    
    try:
        # 读取Excel文件
        print("正在读取Excel文件...")
        df = pd.read_excel(excel_file)
        
        # 检查是否包含PatientID列
        if 'PatientID' not in df.columns:
            print("错误: Excel文件中没有找到'PatientID'列")
            print(f"可用列: {list(df.columns)}")
            return False
        
        print(f"原始表格形状: {df.shape}")
        print(f"PatientID列示例: {df['PatientID'].head().tolist()}")
        
        # 确保PatientID列是字符串类型，便于比较
        df['PatientID'] = df['PatientID'].astype(str)
        
        # 创建表格1：需要重新导出的数据
        print("\n创建表格1: 需要重新导出")
        table1 = df[df['PatientID'].isin(patient_ids_to_filter)].copy()
        print(f"表格1记录数: {len(table1)}")
        
        # 检查是否所有PatientID都找到了
        found_ids = table1['PatientID'].tolist()
        missing_ids = [pid for pid in patient_ids_to_filter if pid not in found_ids]
        
        if missing_ids:
            print(f"警告: 以下PatientID在原始表中未找到: {missing_ids}")
        
        # 创建表格2：剩下的数据
        print("\n创建表格2: 剩下入选")
        table2 = df[~df['PatientID'].isin(patient_ids_to_filter)].copy()
        print(f"表格2记录数: {len(table2)}")
        
        # 使用openpyxl来操作Excel文件，保留原始格式
        print("\n正在写入Excel文件...")
        
        # 加载现有的Excel文件
        book = load_workbook(excel_file)
        
        # 删除已存在的同名工作表（如果存在）
        if '需要重新导出' in book.sheetnames:
            del book['需要重新导出']
        if '剩下入选' in book.sheetnames:
            del book['剩下入选']
        
        # 创建新的工作表
        ws1 = book.create_sheet('需要重新导出')
        ws2 = book.create_sheet('剩下入选')
        
        # 将数据写入工作表
        # 表格1
        for r in dataframe_to_rows(table1, index=False, header=True):
            ws1.append(r)
        
        # 表格2
        for r in dataframe_to_rows(table2, index=False, header=True):
            ws2.append(r)
        
        # 保存文件
        book.save(excel_file)
        
        print("=" * 50)
        print("操作完成!")
        print(f"表格1 '需要重新导出': {len(table1)} 条记录")
        print(f"表格2 '剩下入选': {len(table2)} 条记录")
        
        # 显示一些统计信息
        if len(table1) > 0:
            print(f"\n表格1中的PatientID: {table1['PatientID'].tolist()}")
        
        return True
        
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        return False

def alternative_method():
    """
    备选方法：使用pandas直接处理（可能会丢失格式）
    """
    excel_file = "/media/dell/T7 Shield/nnunet/AllData/ZJYY/珠江需要重新导出10例.xlsx"
    patient_ids_to_filter = [
        "4167356", "4600423", "4605329", "4619739", "4646742", 
        "4673571", "4690443", "4938852", "2901533","4128007"
    ]
    
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_file)
        df['PatientID'] = df['PatientID'].astype(str)
        
        # 创建两个子表格
        table1 = df[df['PatientID'].isin(patient_ids_to_filter)]
        table2 = df[~df['PatientID'].isin(patient_ids_to_filter)]
        
        # 使用ExcelWriter来写入多个工作表
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            table1.to_excel(writer, sheet_name='需要重新导出', index=False)
            table2.to_excel(writer, sheet_name='剩下入选', index=False)
        
        print("备选方法完成!")
        return True
        
    except Exception as e:
        print(f"备选方法错误: {e}")
        return False

def check_excel_structure():
    """
    检查Excel文件的结构
    """
    excel_file = "/media/dell/T7 Shield/nnunet/AllData/ZJYY/珠江需要重新导出10例.xlsx"
    
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_file)
        
        print("Excel文件结构:")
        print(f"行数: {len(df)}")
        print(f"列数: {len(df.columns)}")
        print(f"列名: {list(df.columns)}")
        print(f"前5行数据:")
        print(df.head())
        
        # 检查PatientID列的数据类型和示例
        if 'PatientID' in df.columns:
            print(f"\nPatientID列信息:")
            print(f"数据类型: {df['PatientID'].dtype}")
            print(f"唯一值示例: {df['PatientID'].unique()[:10]}")
            print(f"空值数量: {df['PatientID'].isnull().sum()}")
        
        return True
        
    except Exception as e:
        print(f"检查文件结构时出错: {e}")
        return False

def main():
    """
    主函数
    """
    print("开始处理Excel文件...")
    print("=" * 50)
    
    # 首先检查文件结构
    print("检查Excel文件结构...")
    if not check_excel_structure():
        print("文件结构检查失败，请检查文件路径和格式")
        return
    
    print("\n" + "=" * 50)
    print("开始创建子表格...")
    
    # 使用方法1创建子表格
    success = create_subtables_from_excel()
    
    if not success:
        print("\n方法1失败，尝试备选方法...")
        success = alternative_method()
    
    if success:
        print("\n操作成功完成!")
        print("请打开Excel文件查看新创建的工作表:")
        print("1. '需要重新导出' - 包含指定的PatientID")
        print("2. '剩下入选' - 包含其他的PatientID")
    else:
        print("\n所有方法都失败了，请检查Excel文件格式")

if __name__ == "__main__":
    main()